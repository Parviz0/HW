import openpyxl
from openpyxl.styles import Font, Border, Side

file_names = ['C:\\Users\\User\\Desktop\\Программирование\\Python\\HomeWork11\\Книга1.xlsx']

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

for file in file_names:
    wb = openpyxl.load_workbook(file)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        ws_new.title = sheet_name + "_sorted"
        ws_new.append(["Содержимое файла", file, "и листа", sheet_name, "(отсортированное в порядке убывания):"])
        data = [row for row in sheet.iter_rows(values_only=True)]
        sorted_data = sorted(data, reverse=True)
        for row in sorted_data:
            ws_new.append(row)

font_style = Font(name='Calibri', size=12, bold=True)
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws_new.iter_rows(min_row=1, max_col=ws_new.max_column, max_row=ws_new.max_row):
    for cell in row:
        cell.font = font_style
        cell.border = border_style

wb_new.save('C:\\Users\\User\\Desktop\\Программирование\\Python\\HomeWork11\\Sorted_Kнига1.xlsx')

# Чтобы код заработал скачайте библиотеки pandas и openpyxl и думаю надо будет поменять путь